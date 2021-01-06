import openpyxl as xl

numbers_wb = xl.load_workbook('1. Meal Count INPUT.xlsx')
numbers_sheet = numbers_wb['COUNT']


### EVERY WEEK ENTER MEALS

red_meal = 'Chicken Tagine'
teal_meal = 'Spaghetti and Turkey Meatballs'
green_meal = 'Mexican Lasagna'
yellow_meal = 'Sweet Corn Arancini'
blue_meal = 'Spaghetti and Vegetarian Meatballs'

date = input('DATE: ')


##COLUMN DESIGNATION
date_column = 22
red_column = 22
teal_column = 23
green_column = 24
yellow_column = 25
blue_column = 26

date_cell = numbers_sheet.cell(1, red_column)
date_cell.value = date

meal_column = 4
for row in range(2, numbers_sheet.max_row + 1):
    meal_cell = numbers_sheet.cell(row, meal_column)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal in meals:
        meals = meals.replace(blue_meal, 'blue')
        meals = meals.replace(yellow_meal, 'yellow')
        meals = meals.replace(green_meal, 'green')
        meals = meals.replace(teal_meal, 'teal')
        meals = meals.replace(red_meal, 'red')
        meal_cell.value = meals
    else:
        meal_column = 4

numbers_wb.save('2. Meal Count DAILY.xlsx')
