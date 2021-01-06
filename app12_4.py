import openpyxl as xl


###COPY MEAL POT COLORS HERE


red_meal = 'Smoked Brisket & Potato Latkes'
teal_meal = 'Chicken Fajitas'
green_meal = 'Potato Gnocchi'
yellow_meal = 'Vegetable Potstickers '
blue_meal = 'Vegetable Fajitas'



red_meal_input = input('RED POT: ')
teal_meal_input = input('TEAL POT: ')
green_meal_input = input('GREEN POT: ')
yellow_meal_input = input('YELLOW POT: ')
blue_meal_input = input('BLUE POT: ')




wb = xl.load_workbook('5. Shipping Report Input.xlsx')
sheet = wb['ShipRep_GF']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']
alpha_list_number = 0
addon_list_number = 0
row = 2
for row in range(2, sheet.max_row + 1):
    previous_row = row - 1
    reference2 = sheet.cell(row, 5)
    previous_reference2 = sheet.cell(previous_row, 5)
    number_of_smoothies_cell = sheet.cell(row, 6)
    number_of_smoothies = number_of_smoothies_cell.value
    number_of_cookies = sheet.cell(row, 15)
    alpha_cell = sheet.cell(row, 26)
    if row == 2 and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number]
    elif reference2.value == previous_reference2.value and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number]
    elif reference2.value != previous_reference2.value and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number + 1]
        alpha_list_number = alpha_list_number + 1
    else:
        alpha_cell = addon_list[addon_list_number]
        addon_list_number = addon_list_number + 1
    alpha_cell_update = sheet.cell(row, 26)
    alpha_cell_update.value = alpha_cell

###REG ALPHAS###
sheet = wb['ShipRep_REG']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

alpha_list_number = 0
addon_list_number = 0
row = 2
for row in range(2, sheet.max_row + 1):
    previous_row = row - 1
    reference2 = sheet.cell(row, 5)
    previous_reference2 = sheet.cell(previous_row, 5)
    number_of_smoothies_cell = sheet.cell(row, 6)
    number_of_smoothies = number_of_smoothies_cell.value
    number_of_cookies = sheet.cell(row, 15)
    alpha_cell = sheet.cell(row, 26)
    if row == 2 and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number]
    elif reference2.value == previous_reference2.value and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number]
    elif reference2.value != previous_reference2.value and number_of_smoothies_cell.value == 0 and number_of_cookies.value == 0:
        alpha_cell = alpha_list[alpha_list_number + 1]
        alpha_list_number = alpha_list_number + 1
    else:
        alpha_cell = addon_list[addon_list_number]
        addon_list_number = addon_list_number + 1
    alpha_cell_update = sheet.cell(row, 26)
    alpha_cell_update.value = alpha_cell


sheet = wb['ShipRep_GF']
list_sheet = wb['GF_LIST']
row = 2
list_sheet_row = 4
for row in range(2, sheet.max_row + 1):
    previous_row = row - 1
    alpha_cell = sheet.cell(row, 26)
    previous_reference2 = sheet.cell(previous_row, 5)
    previous_alpha_cell = sheet.cell(row - 1, 26)
    list_sheet_alpha = list_sheet.cell(list_sheet_row, 1)
    if alpha_cell.value != previous_alpha_cell.value:
        list_sheet_cell = alpha_cell
        list_sheet_alpha.value = list_sheet_cell.value
        row + 1
        row = row + 1
        list_sheet_row = list_sheet_row + 1
    else:
        row = row + 1
wb.save('00. DO NOT USE.xlsx')

sheet = wb['ShipRep_REG']
list_sheet = wb['REG_LIST']
row = 2
list_sheet_row = 4
for row in range(2, sheet.max_row + 1):
    previous_row = row - 1
    alpha_cell = sheet.cell(row, 26)
    previous_reference2 = sheet.cell(previous_row, 5)
    previous_alpha_cell = sheet.cell(row - 1, 26)
    list_sheet_alpha = list_sheet.cell(list_sheet_row, 1)
    if alpha_cell.value != previous_alpha_cell.value:
        list_sheet_cell = alpha_cell
        list_sheet_alpha.value = list_sheet_cell.value
        row + 1
        row = row + 1
        list_sheet_row = list_sheet_row + 1
    else:
        row = row + 1
wb.save('00. DO NOT USE.xlsx')


###SHIPPING INPUT NUMBER CHANGE
###MEAL CHANGE INPUT
delivery_routes = ['Mid-City', 'LA46', 'SEALED-BAG','Beverly Hills', 'Santa Monica', 'Valley', 'Long Beach', 'Pasadena', 'Los Angeles', 'Los Angeles 04', 'Orange County', 'Brentwood','Manhattan Beach', 'Studio City','Studio City 2', 'Valley2','Hollywood', 'Hollywood 2','Culver City','Woodland Hills']

sheet = wb['ShipRep_GF']
row = 2
list_sheet_row = 4
meal_column = 5
delivery_column = 25
for row in range(2, sheet.max_row + 1):
    meal_cell = sheet.cell(row, meal_column)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal or 'Vegetarian Pizza' or 'Classic Burger & Fries' or 'Classic Turkey Burger & Fries' or 'Classic Veggie Burger & Fries' or 'Beef Tacos' or 'Turkey Tacos' or 'Veggie Tacos' in meals:
        meals = meals.replace(blue_meal, blue_meal_input)
        meals = meals.replace(yellow_meal, yellow_meal_input)
        meals = meals.replace(green_meal, green_meal_input)
        meals = meals.replace(teal_meal, teal_meal_input)
        meals = meals.replace(red_meal, red_meal_input)
        meals = meals.replace('Vegetarian Pizza', 'V-Pizza')
        meals = meals.replace('Classic Burger & Fries', 'B-Burg')
        meals = meals.replace('Classic Turkey Burger & Fries', 'T-Burg')
        meals = meals.replace('Classic Veggie Burger & Fries', 'V-Burg')
        meals = meals.replace('Beef Tacos', 'B-Tacos')
        meals = meals.replace('Turkey Tacos', 'T-Tacos')
        meals = meals.replace('Veggie Tacos', 'V-Tacos')
        meal_cell.value = meals
for row in range(2, sheet.max_row + 1):
    bag_cell = sheet.cell(row, delivery_column)
    bags = str(bag_cell.value)
    if bags in delivery_routes:
        bags = 1
        bag_cell.value = bags
    elif bag_cell.value is None:
        bags = 0
        bag_cell.value = bags




sheet = wb['ShipRep_REG']
row = 2
list_sheet_row = 4
meal_column = 5
for row in range(2, sheet.max_row + 1):
    meal_cell = sheet.cell(row, meal_column)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal or 'Vegetarian Pizza' or 'Classic Burger & Fries' or 'Classic Turkey Burger & Fries' or 'Classic Veggie Burger & Fries' or 'Beef Tacos' or 'Turkey Tacos' or 'Veggie Tacos' in meals:
        meals = meals.replace(blue_meal, blue_meal_input)
        meals = meals.replace(yellow_meal, yellow_meal_input)
        meals = meals.replace(green_meal, green_meal_input)
        meals = meals.replace(teal_meal, teal_meal_input)
        meals = meals.replace(red_meal, red_meal_input)
        meals = meals.replace('Vegetarian Pizza', 'V-Pizza')
        meals = meals.replace('Classic Burger & Fries', 'B-Burg')
        meals = meals.replace('Classic Turkey Burger & Fries', 'T-Burg')
        meals = meals.replace('Classic Veggie Burger & Fries', 'V-Burg')
        meals = meals.replace('Beef Tacos', 'B-Tacos')
        meals = meals.replace('Turkey Tacos', 'T-Tacos')
        meals = meals.replace('Veggie Tacos', 'V-Tacos')
        meal_cell.value = meals


for row in range(2, sheet.max_row + 1):
    bag_cell = sheet.cell(row, delivery_column)
    bags = str(bag_cell.value)
    if bags in delivery_routes:
        bags = 1
        bag_cell.value = bags
    elif bag_cell.value is None:
        bags = 0
        bag_cell.value = bags


wb.save('00. DO NOT USE.xlsx')



####MEAL NAMES
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_GF']
list_sheet = wb['GF_LIST']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']

reference = 4
reference_2 = 5
alpha_column = 26
meals = 5
sheet_row = 4
for row in range(2, sheet.max_row + 1):
    sheet_meal_cell = sheet.cell(row, reference_2)
    sheet_previous_meal_cell = sheet.cell(row - 1, reference_2)
    sheet_alpha_cell = sheet.cell(row, alpha_column)
    list_sheet_meal_cell = list_sheet.cell(sheet_row, meals)
    alpha = sheet_alpha_cell.value
    list_alpha = list_sheet.cell(sheet_row, 1)
    reference_cell = sheet.cell(row, reference)
    list_sheet_kids = list_sheet.cell(sheet_row, 4)
    if alpha in alpha_list and alpha == list_alpha.value:
            sheet_row = sheet_row + 1
            list_sheet_meal_cell.value = sheet_meal_cell.value
            list_sheet_kids.value = reference_cell.value
    elif alpha in alpha_list and sheet_meal_cell.value != list_alpha.value:
            sheet_row = sheet_row + 0
            list_sheet_meal_cell.value = sheet_meal_cell.value
            list_sheet_kids.value = reference_cell.value
    else:
        sheet_row = sheet_row + 1
        list_sheet_meal_cell.value = sheet_meal_cell.value
        list_sheet_kids.value = reference_cell.value
wb.save('00. DO NOT USE.xlsx')



###BAG BOX COUNT
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_GF']
list_sheet = wb['GF_LIST']

list_row = 4
number_of_bags = 0
number_of_boxes = 0
alpha_column = 26
bag_column = 25

wb.save('00. DO NOT USE.xlsx')



###BAG BOX COUNT
sheet = wb['ShipRep_REG']
list_sheet = wb['REG_LIST']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

reference = 4
reference_2 = 5
alpha_column = 26
meals = 5
sheet_row = 4
for row in range(2, sheet.max_row + 1):
    sheet_meal_cell = sheet.cell(row, reference_2)
    sheet_previous_meal_cell = sheet.cell(row - 1, reference_2)
    sheet_alpha_cell = sheet.cell(row, alpha_column)
    list_sheet_meal_cell = list_sheet.cell(sheet_row, meals)
    alpha = sheet_alpha_cell.value
    list_alpha = list_sheet.cell(sheet_row, 1)
    reference_cell = sheet.cell(row, reference)
    list_sheet_kids = list_sheet.cell(sheet_row, 4)
    if alpha in alpha_list and alpha == list_alpha.value:
            sheet_row = sheet_row + 1
            list_sheet_meal_cell.value = sheet_meal_cell.value
            list_sheet_kids.value = reference_cell.value
    elif alpha in alpha_list and sheet_meal_cell.value != list_alpha.value:
            sheet_row = sheet_row + 0
            list_sheet_meal_cell.value = sheet_meal_cell.value
            list_sheet_kids.value = reference_cell.value
    else:
        sheet_row = sheet_row + 1
        list_sheet_meal_cell.value = sheet_meal_cell.value
        list_sheet_kids.value = reference_cell.value
wb.save('00. DO NOT USE.xlsx')


###BAG BOX COUNT
import openpyxl as xl
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_GF']
list_sheet = wb['GF_LIST']

for row in range(2, sheet.max_row +1):
    delivery_cell = sheet.cell(row, 25)
    bag_cell = sheet.cell(row, 27)
    if delivery_cell.value == 1:
        bag_cell.value = 0
    elif delivery_cell.value == 0:
        bag_cell.value = 1

wb.save('00. DO NOT USE.xlsx')
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_REG']
list_sheet = wb['REG_LIST']

for row in range(2, sheet.max_row +1):
    delivery_cell = sheet.cell(row, 25)
    bag_cell = sheet.cell(row, 27)
    if delivery_cell.value == 1:
        bag_cell.value = 0
    elif delivery_cell.value == 0:
        bag_cell.value = 1
wb.save('00. DO NOT USE.xlsx')




###FIX ALL THIS
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_GF']
list_sheet = wb['GF_LIST']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']

bag_count = 0
box_count = 0
for list_row in range(4, list_sheet.max_row + 1):
    list_alpha = list_sheet.cell(list_row, 1)
    list_bag = list_sheet.cell(list_row, 2)
    list_box = list_sheet.cell(list_row, 3)
    if list_alpha.value in alpha_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_bag = sheet.cell(row, 25)
            sheet_box = sheet.cell(row, 27)
            if list_alpha.value == sheet_alpha.value:
                if sheet_bag.value == 1 and sheet_box.value == 0:
                    bag_count = bag_count + 1
                    list_bag.value = bag_count
                    list_box.value = box_count
                elif sheet_bag.value == 0 and sheet_box.value == 1:
                    box_count = box_count + 1
                    list_bag.value = bag_count
                    list_box.value = box_count
            else:
                bag_count = 0
                box_count = 0
    elif list_alpha.value in addon_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_bag = sheet.cell(row, 25)
            sheet_box = sheet.cell(row, 27)
            if list_alpha.value == sheet_alpha.value:
                if sheet_bag.value == 1 and sheet_box.value == 0:
                    list_bag.value = 1
                    list_box.value = 0
                    bag_count = 0
                    box_count = 0
                elif sheet_bag.value == 0 and sheet_box.value == 1:
                    list_bag.value = 0
                    list_box.value = 1
                    bag_count = 0
                    box_count = 0
            else:
                bag_count = 0
                box_count = 0

###2 day ice
ice_count = 0
for list_row in range(4, list_sheet.max_row + 1):
    list_alpha = list_sheet.cell(list_row, 1)
    list_ice = list_sheet.cell(list_row, 6)
    if list_alpha.value in alpha_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_ice = sheet.cell(row, 24)
            if list_alpha.value == sheet_alpha.value:
                if sheet_ice.value == '2 Day':
                    ice_count = ice_count + 1
                    list_ice.value = ice_count
                else:
                    ice_count = 0
            else:
                ice_count = 0
    elif list_alpha.value in addon_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_ice = sheet.cell(row, 24)
            if list_alpha.value == sheet_alpha.value:
                if sheet_ice.value == '2 Day':
                    list_ice.value = 1
                elif sheet_ice.value == '1 Day' or sheet_ice.value == '1 Day UPS':
                    list_ice.value = ''
            else:
                ice_count = 0
wb.save('00. DO NOT USE.xlsx')
wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_REG']
list_sheet = wb['REG_LIST']
alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

bag_count = 0
box_count = 0
for list_row in range(4, list_sheet.max_row + 1):
    list_alpha = list_sheet.cell(list_row, 1)
    list_bag = list_sheet.cell(list_row, 2)
    list_box = list_sheet.cell(list_row, 3)
    if list_alpha.value in alpha_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_bag = sheet.cell(row, 25)
            sheet_box = sheet.cell(row, 27)
            if list_alpha.value == sheet_alpha.value:
                if sheet_bag.value == 1 and sheet_box.value == 0:
                    bag_count = bag_count + 1
                    list_bag.value = bag_count
                    list_box.value = box_count
                elif sheet_bag.value == 0 and sheet_box.value == 1:
                    box_count = box_count + 1
                    list_bag.value = bag_count
                    list_box.value = box_count
            else:
                bag_count = 0
                box_count = 0
    elif list_alpha.value in addon_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_bag = sheet.cell(row, 25)
            sheet_box = sheet.cell(row, 27)
            if list_alpha.value == sheet_alpha.value:
                if sheet_bag.value == 1 and sheet_box.value == 0:
                    list_bag.value = 1
                    list_box.value = 0
                    bag_count = 0
                    box_count = 0
                elif sheet_bag.value == 0 and sheet_box.value == 1:
                    list_bag.value = 0
                    list_box.value = 1
                    bag_count = 0
                    box_count = 0
            else:
                bag_count = 0
                box_count = 0

###2 day ice
ice_count = 0
for list_row in range(4, list_sheet.max_row + 1):
    list_alpha = list_sheet.cell(list_row, 1)
    list_ice = list_sheet.cell(list_row, 6)
    if list_alpha.value in alpha_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_ice = sheet.cell(row, 24)
            if list_alpha.value == sheet_alpha.value:
                if sheet_ice.value == '2 Day':
                    ice_count = ice_count + 1
                    list_ice.value = ice_count
                else:
                    ice_count = 0
            else:
                ice_count = 0
    elif list_alpha.value in addon_list:
        for row in range(2, sheet.max_row + 1):
            sheet_alpha = sheet.cell(row, 26)
            sheet_ice = sheet.cell(row, 24)
            if list_alpha.value == sheet_alpha.value:
                if sheet_ice.value == '2 Day':
                    list_ice.value = 1
                elif sheet_ice.value == '1 Day' or sheet_ice.value == '1 Day UPS':
                    list_ice.value = ''
            else:
                ice_count = 0

wb.save('00. DO NOT USE.xlsx')

wb = xl.load_workbook('00. DO NOT USE.xlsx')
list_sheet = wb['GF_LIST']

for row in range(4, list_sheet.max_row + 1):
    zero_bag = list_sheet.cell(row, 2)
    if zero_bag.value == 0:
        zero_bag.value = ''
    else:
        zero_bag.value = zero_bag.value

wb.save('00. DO NOT USE.xlsx')

wb = xl.load_workbook('00. DO NOT USE.xlsx')
list_sheet = wb['REG_LIST']

for row in range(4, list_sheet.max_row + 1):
    zero_bag = list_sheet.cell(row, 2)
    if zero_bag.value == 0:
        zero_bag.value = ''
    else:
        zero_bag.value = zero_bag.value


wb.save('00. DO NOT USE.xlsx')

wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_VIP']
for row in range(2, sheet.max_row + 1):
    meal_cell = sheet.cell(row, 6)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal or 'Vegetarian Pizza' or 'Classic Burger & Fries' or 'Classic Turkey Burger & Fries' or 'Classic Veggie Burger & Fries' or 'Beef Tacos' or 'Turkey Tacos' or 'Veggie Tacos' in meals:
        meals = meals.replace(blue_meal, blue_meal_input)
        meals = meals.replace(yellow_meal, yellow_meal_input)
        meals = meals.replace(green_meal, green_meal_input)
        meals = meals.replace(teal_meal, teal_meal_input)
        meals = meals.replace(red_meal, red_meal_input)
        meals = meals.replace('Vegetarian Pizza', 'V-Pizza')
        meals = meals.replace('Classic Burger & Fries', 'B-Burg')
        meals = meals.replace('Classic Turkey Burger & Fries', 'T-Burg')
        meals = meals.replace('Classic Veggie Burger & Fries', 'V-Burg')
        meals = meals.replace('Beef Tacos', 'B-Tacos')
        meals = meals.replace('Turkey Tacos', 'T-Tacos')
        meals = meals.replace('Veggie Tacos', 'V-Tacos')
        meal_cell.value = meals


wb.save('00. DO NOT USE.xlsx')

wb = xl.load_workbook('00. DO NOT USE.xlsx')
sheet = wb['ShipRep_VIP']
ice_column = 25
for row in range(2, sheet.max_row + 1):
    ice_cell = sheet.cell(row, ice_column)
    if ice_cell.value == "1 Day":
        ice_cell.value = ""
    elif ice_cell.value == "1 Day UPS":
        ice_cell.value = ""
    elif ice_cell.value == "2 Day":
        ice_cell.value = "2 Day"

sheet = wb['ShipRep_GF']
ice_column = 24
for row in range(2, sheet.max_row + 1):
    ice_cell = sheet.cell(row, ice_column)
    if ice_cell.value == "1 Day":
        ice_cell.value = ""
    elif ice_cell.value == "1 Day UPS":
        ice_cell.value = ""
    elif ice_cell.value == "2 Day":
        ice_cell.value = "2 Day"

sheet = wb['ShipRep_REG']
ice_column = 24
for row in range(2, sheet.max_row + 1):
    ice_cell = sheet.cell(row, ice_column)
    if ice_cell.value == "1 Day":
        ice_cell.value = ""
    elif ice_cell.value == "1 Day UPS":
        ice_cell.value = ""
    elif ice_cell.value == "2 Day":
        ice_cell.value = "2 Day"

wb.save('00. DO NOT USE.xlsx')

import openpyxl
wb = openpyxl.load_workbook('00. DO NOT USE.xlsx')
from openpyxl.styles import Font, Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.cell import Cell
from copy import copy, deepcopy

sheet = wb['GF_LIST']
fourk_cell = sheet.cell(1, 8)
fourk_addon_cell = sheet.cell(2,8)
threek_cell = sheet.cell(3, 8)
threek_addon_cell = sheet.cell(4,8)
twok_cell = sheet.cell(5, 8)
twok_addon_cell = sheet.cell(6,8)
onek_cell = sheet.cell(7, 8)
onek_addon_cell = sheet.cell(8,8)
zerok_cell = sheet.cell(9, 8)
zerok_addon_cell = sheet.cell(10,8)
three_meal_cell = sheet.cell(11.,8)
if fourk_cell.has_style:
    if 'FOURk_style' not in wb.named_styles:
        FOURk_style = NamedStyle(name='FOURk_style')
        FOURk_style.font = copy(fourk_cell.font)
        FOURk_style.border = copy(fourk_cell.border)
        FOURk_style.fill = copy(fourk_cell.fill)
        FOURk_style.number_format = copy(fourk_cell.number_format)
        FOURk_style.protection = copy(fourk_cell.protection)
        FOURk_style.alignment = copy(fourk_cell.alignment)
        wb.add_named_style(FOURk_style)
if fourk_addon_cell.has_style:
    if 'FOURk_ADDON_style' not in wb.named_styles:
        FOURk_ADDON_style = NamedStyle(name='FOURk_ADDON_style')
        FOURk_ADDON_style.font = copy(fourk_addon_cell.font)
        FOURk_ADDON_style.border = copy(fourk_addon_cell.border)
        FOURk_ADDON_style.fill = copy(fourk_addon_cell.fill)
        FOURk_ADDON_style.number_format = copy(fourk_addon_cell.number_format)
        FOURk_ADDON_style.protection = copy(fourk_addon_cell.protection)
        FOURk_ADDON_style.alignment = copy(fourk_addon_cell.alignment)
        wb.add_named_style(FOURk_ADDON_style)
if threek_cell.has_style:
    if 'THREEk_style' not in wb.named_styles:
        THREEk_style = NamedStyle(name='THREEk_style')
        THREEk_style.font = copy(threek_cell.font)
        THREEk_style.border = copy(threek_cell.border)
        THREEk_style.fill = copy(threek_cell.fill)
        THREEk_style.number_format = copy(threek_cell.number_format)
        THREEk_style.protection = copy(threek_cell.protection)
        THREEk_style.alignment = copy(threek_cell.alignment)
        wb.add_named_style(THREEk_style)
if threek_addon_cell.has_style:
    if 'THREEk_ADDON_style' not in wb.named_styles:
        THREEk_ADDON_style = NamedStyle(name='THREEk_ADDON_style')
        THREEk_ADDON_style.font = copy(threek_addon_cell.font)
        THREEk_ADDON_style.border = copy(threek_addon_cell.border)
        THREEk_ADDON_style.fill = copy(threek_addon_cell.fill)
        THREEk_ADDON_style.number_format = copy(threek_addon_cell.number_format)
        THREEk_ADDON_style.protection = copy(threek_addon_cell.protection)
        THREEk_ADDON_style.alignment = copy(threek_addon_cell.alignment)
        wb.add_named_style(THREEk_ADDON_style)
if twok_cell.has_style:
    if 'TWOk_style' not in wb.named_styles:
        TWOk_style = NamedStyle(name='TWOk_style')
        TWOk_style.font = copy(twok_cell.font)
        TWOk_style.border = copy(twok_cell.border)
        TWOk_style.fill = copy(twok_cell.fill)
        TWOk_style.number_format = copy(twok_cell.number_format)
        TWOk_style.protection = copy(twok_cell.protection)
        TWOk_style.alignment = copy(twok_cell.alignment)
        wb.add_named_style(TWOk_style)
if twok_addon_cell.has_style:
    if 'TWOk_ADDON_style' not in wb.named_styles:
        TWOk_ADDON_style = NamedStyle(name='TWOk_ADDON_style')
        TWOk_ADDON_style.font = copy(twok_addon_cell.font)
        TWOk_ADDON_style.border = copy(twok_addon_cell.border)
        TWOk_ADDON_style.fill = copy(twok_addon_cell.fill)
        TWOk_ADDON_style.number_format = copy(twok_addon_cell.number_format)
        TWOk_ADDON_style.protection = copy(twok_addon_cell.protection)
        TWOk_ADDON_style.alignment = copy(twok_addon_cell.alignment)
        wb.add_named_style(TWOk_ADDON_style)
if onek_cell.has_style:
    if 'ONEk_style' not in wb.named_styles:
        ONEk_style = NamedStyle(name='ONEk_style')
        ONEk_style.font = copy(onek_cell.font)
        ONEk_style.border = copy(onek_cell.border)
        ONEk_style.fill = copy(onek_cell.fill)
        ONEk_style.number_format = copy(onek_cell.number_format)
        ONEk_style.protection = copy(onek_cell.protection)
        ONEk_style.alignment = copy(onek_cell.alignment)
        wb.add_named_style(ONEk_style)
if onek_addon_cell.has_style:
    if 'ONEk_ADDON_style' not in wb.named_styles:
        ONEk_ADDON_style = NamedStyle(name='ONEk_ADDON_style')
        ONEk_ADDON_style.font = copy(onek_addon_cell.font)
        ONEk_ADDON_style.border = copy(onek_addon_cell.border)
        ONEk_ADDON_style.fill = copy(onek_addon_cell.fill)
        ONEk_ADDON_style.number_format = copy(onek_addon_cell.number_format)
        ONEk_ADDON_style.protection = copy(onek_addon_cell.protection)
        ONEk_ADDON_style.alignment = copy(onek_addon_cell.alignment)
        wb.add_named_style(ONEk_ADDON_style)
if zerok_cell.has_style:
    if 'ZEROk_style' not in wb.named_styles:
        ZEROk_style = NamedStyle(name='ZEROk_style')
        ZEROk_style.font = copy(zerok_cell.font)
        ZEROk_style.border = copy(zerok_cell.border)
        ZEROk_style.fill = copy(zerok_cell.fill)
        ZEROk_style.number_format = copy(zerok_cell.number_format)
        ZEROk_style.protection = copy(zerok_cell.protection)
        ZEROk_style.alignment = copy(zerok_cell.alignment)
        wb.add_named_style(ZEROk_style)
if zerok_addon_cell.has_style:
    if 'ZEROk_ADDON_style' not in wb.named_styles:
        ZEROk_ADDON_style = NamedStyle(name='ZEROk_ADDON_style')
        ZEROk_ADDON_style.font = copy(zerok_addon_cell.font)
        ZEROk_ADDON_style.border = copy(zerok_addon_cell.border)
        ZEROk_ADDON_style.fill = copy(zerok_addon_cell.fill)
        ZEROk_ADDON_style.number_format = copy(zerok_addon_cell.number_format)
        ZEROk_ADDON_style.protection = copy(zerok_addon_cell.protection)
        ZEROk_ADDON_style.alignment = copy(zerok_addon_cell.alignment)
        wb.add_named_style(ZEROk_ADDON_style)
if three_meal_cell.has_style:
    if 'three_meal_style' not in wb.named_styles:
        three_meal_style = NamedStyle(name='three_meal_style')
        three_meal_style.font = copy(three_meal_cell.font)
        three_meal_style.border = copy(three_meal_cell.border)
        three_meal_style.fill = copy(three_meal_cell.fill)
        three_meal_style.number_format = copy(three_meal_cell.number_format)
        three_meal_style.protection = copy(three_meal_cell.protection)
        three_meal_style.alignment = copy(three_meal_cell.alignment)
        wb.add_named_style(three_meal_style)

###  GF

alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']

alpha_column = 1
meals_column = 5
kids_column = 4
bag_column = 2
box_column = 3
meal_column = 5
ice_column = 6

for row in range(4, sheet.max_row + 1):
    alpha = sheet.cell(row, alpha_column)
    kids = sheet.cell(row, kids_column)
    if kids.value == '4' and alpha.value in alpha_list:
        for col in range(1, 7):
            Cell = sheet.cell(row, col)
            Cell.style = FOURk_style
    elif kids.value == '4' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = FOURk_ADDON_style
    elif kids.value == '3' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = THREEk_style
    elif kids.value == '3' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = THREEk_ADDON_style
    elif kids.value == '2' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = TWOk_style
    elif kids.value == '2' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = TWOk_ADDON_style
    elif kids.value == '1' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = ONEk_style
    elif kids.value == '1' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = ONEk_ADDON_style
    elif kids.value == '0' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = ZEROk_style
    elif kids.value == '0' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = sheet.cell(row, col)
            Cell.style = ZEROk_ADDON_style
wb.save('6. Shipping Report Packlist OUTPUT.xlsx')


alpha_list = ['A', 'B', 'C', 'D', 'E','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','2A','2B','2C','2D','2E','2F','2G','2H',
              '2I','2J','2K','2L','2M','2N','2O','2P','2Q','2R','2S','2T','2U','2V','2W','2X','2Y','2Z','3A','3B','3C','3D','3E','3F','3G','3H',
              '3I','3J','3K','3L','3M','3N','3O','3P','3Q','3R','3S','3T','3U','3V','3W','3X','3Y','3Z','4A','4B','4C','4D','4E','4F','4G','4H',
              '4I','4J','4K','4L','4M','4N','4O','4P','4Q','4R','4S','4T','4U','4V','4W','4X','4Y','4Z','5A','5B','5C','5D','5E','5F','5G','5H',
              '5I','5J','5K','5L','5M','5N','5O','5P','5Q','5R','5S','5T','5U','5V','5W','5X','5Y','5Z','6A','6B','6C','6D','6E','6F','6G','6H',
              '6I','6J','6K','6L','6M','6N','6O','6P','6Q','6R','6S','6T','6U','6V','6W','6X','6Y','6Z','7A','7B','7C','7D','7E','7F','7G','7H',
              '7I','7J','7K','7L','7M','7N','7O','7P','7Q','7R','7S','7T','7U','7V','7W','7X','7Y','7Z','8A','8B','8C','8D','8E','8F','8G','8H',
              '8I','8J','8K','8L','8M','8N','8O','8P','8Q','8R','8S','8T','8U','8V','8W','8X','8Y','8Z','9A','9B','9C','9D','9E','9F','9G','9H',
              '9I','9J','9K','9L','9M','9N','9O','9P','9Q','9R','9S','9T','9U','9V','9W','9X','9Y','9Z','10A','10B','10C','10D','10E','10F','10G','10H',
              '10I','10J','10K','10L','10M','10N','10O','10P','10Q','10R','10S','10T','10U','10V','10W','10X','10Y','10Z','11A','11B','11C','11D','11E','11F','11G','11H',
              '11I','11J','11K','11L','11M','11N','11O','11P','11Q','11R','11S','11T','11U','11V','11W','11X','11Y','11Z']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

alpha_column = 1
meals_column = 5
kids_column = 4
bag_column = 2
box_column = 3
meal_column = 5
ice_column = 6

reg_sheet = wb['REG_LIST']
for row in range(4, reg_sheet.max_row + 1):
    alpha = reg_sheet.cell(row, alpha_column)
    kids = reg_sheet.cell(row, kids_column)
    if kids.value == '4' and alpha.value in alpha_list:
        for col in range(1, 7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = FOURk_style
    elif kids.value == '4' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = FOURk_ADDON_style
    elif kids.value == '3' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = THREEk_style
    elif kids.value == '3' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = THREEk_ADDON_style
    elif kids.value == '2' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = TWOk_style
    elif kids.value == '2' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = TWOk_ADDON_style
    elif kids.value == '1' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = ONEk_style
    elif kids.value == '1' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = ONEk_ADDON_style
    elif kids.value == '0' and alpha.value in alpha_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = ZEROk_style
    elif kids.value == '0' and alpha.value in addon_list:
        for col in range(1,7):
            Cell = reg_sheet.cell(row, col)
            Cell.style = ZEROk_ADDON_style


wb.save('6. Shipping Report Packlist OUTPUT.xlsx')

for row in range(4, sheet.max_row + 1):
    reference = sheet.cell(row, 5)
    meals = reference.value
    number_of_meals = meals.count(',')
    if number_of_meals == 3:
        reference.style = three_meal_style
    else:
        number_of_meals = 0

wb.save('6. Shipping Report Packlist OUTPUT.xlsx')

for row in range(4, reg_sheet.max_row + 1):
    reference = reg_sheet.cell(row, 5)
    meals = reference.value
    number_of_meals = meals.count(',')
    if number_of_meals == 3:
        reference.style = three_meal_style
    else:
        number_of_meals = 0

wb.save('6. Shipping Report Packlist OUTPUT.xlsx')

###CHECKPOINT
import openpyxl
wb = openpyxl.load_workbook('6. Shipping Report Packlist OUTPUT.xlsx')
from openpyxl.styles import Font, Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.cell import Cell
from copy import copy, deepcopy



###DONT COPY
red_meal = 'Italian Braised Chicken'
teal_meal = 'Chicken Lo Mein'
green_meal = 'Sweet Corn Ravioli'
yellow_meal = 'Potato Spinach Dosas'
blue_meal = 'Vegetable Lo Mein'


####DONT COPY

##TOTAL NUMBERS

sheet = wb['TOTAL_NUMBERS']

for row in range(2, sheet.max_row + 1):
    gf_cell = sheet.cell(row, 3)
    if gf_cell.value == "X" or gf_cell.value == "x":
        gf_cell.value = "Gluten Free"

meal_column = 6
for row in range(2, sheet.max_row + 1):
    meal_cell = sheet.cell(row, meal_column)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal in meals:
        meals = meals.replace(blue_meal, 'blue')
        meals = meals.replace(yellow_meal, 'yellow')
        meals = meals.replace(green_meal, 'green')
        meals = meals.replace(teal_meal, 'teal')
        meals = meals.replace(red_meal, 'red')
        meal_cell.value = meals
    else:
        meal_column = 6



###REG SMOOTHIES
reg_sheet = wb['ShipRep_REG']
list_sheet = wb['REG_ADDONS']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

alpha_column = 26
name_column = 1
kids_column = 4
packlist_row = 4
smoothie_column = 6
antiox_column = 8
bigred_column = 9
blue_column = 10
boost_column = 11
green_column = 12
pbj_column = 13
vitc_column = 14
cookies_column = 15
break_column = 17
chip_column = 18
chipgf_column = 19
or_column = 20
orgf_column = 21
sugar_column = 22
sugargf_column = 23

for row in range(2, reg_sheet.max_row +1):
    alpha_cell = reg_sheet.cell(row, alpha_column)
    name_cell = reg_sheet.cell(row, name_column)
    kids_cell = reg_sheet.cell(row,kids_column)
    smoothie_cell = reg_sheet.cell(row,smoothie_column)
    antiox_cell = reg_sheet.cell(row, antiox_column)
    bigred_cell = reg_sheet.cell(row, bigred_column)
    blue_cell = reg_sheet.cell(row, blue_column)
    boost_cell = reg_sheet.cell(row, boost_column)
    green_cell = reg_sheet.cell(row, green_column)
    pbj_cell = reg_sheet.cell(row, pbj_column)
    vitc_cell = reg_sheet.cell(row,vitc_column)
    cookies_cell = reg_sheet.cell(row, cookies_column)
    break_cell = reg_sheet.cell(row, break_column)
    chip_cell = reg_sheet.cell(row, chip_column)
    chipgf_cell = reg_sheet.cell(row, chipgf_column)
    or_cell = reg_sheet.cell(row, or_column)
    orgf_cell = reg_sheet.cell(row, orgf_column)
    sugar_cell = reg_sheet.cell(row,sugar_column)
    sugargf_cell = reg_sheet.cell(row,sugargf_column)
    alpha = alpha_cell.value
    name = name_cell.value
    kids = kids_cell.value
    if alpha_cell.value in addon_list:
        pack_alpha = list_sheet.cell(packlist_row, 1)
        pack_kids = list_sheet.cell(packlist_row, 2)
        pack_name = list_sheet.cell(packlist_row, 3)
        pack_smoothies = list_sheet.cell(packlist_row, 4)
        pack_antiox = list_sheet.cell(packlist_row, 5)
        pack_bigred = list_sheet.cell(packlist_row, 6)
        pack_blue = list_sheet.cell(packlist_row, 7)
        pack_boost = list_sheet.cell(packlist_row, 8)
        pack_green = list_sheet.cell(packlist_row, 9)
        pack_pbj = list_sheet.cell(packlist_row, 10)
        pack_vitc = list_sheet.cell(packlist_row, 11)
        pack_cookie = list_sheet.cell(packlist_row, 12)
        pack_break = list_sheet.cell(packlist_row, 13)
        pack_chip = list_sheet.cell(packlist_row, 14)
        pack_chipgf = list_sheet.cell(packlist_row, 15)
        pack_or = list_sheet.cell(packlist_row, 16)
        pack_orgf = list_sheet.cell(packlist_row, 17)
        pack_sugar = list_sheet.cell(packlist_row, 18)
        pack_sugargf = list_sheet.cell(packlist_row, 19)
        pack_alpha.value = alpha
        pack_name.value = name
        pack_kids.value = kids
        pack_smoothies.value = smoothie_cell.value
        pack_antiox.value = antiox_cell.value
        pack_bigred.value = bigred_cell.value
        pack_blue.value = blue_cell.value
        pack_boost.value = boost_cell.value
        pack_green.value = green_cell.value
        pack_pbj.value = pbj_cell.value
        pack_vitc.value = vitc_cell.value
        pack_cookie.value = cookies_cell.value
        pack_break.value = break_cell.value
        pack_chip.value = chip_cell.value
        pack_chipgf.value = chipgf_cell.value
        pack_or.value = or_cell.value
        pack_orgf.value = orgf_cell.value
        pack_sugar.value = sugar_cell.value
        pack_sugargf.value = sugargf_cell.value
        packlist_row = packlist_row + 1
wb.save('000. DONOTUSE.xlsx')
wb = openpyxl.load_workbook('000. DONOTUSE.xlsx')
##REG SMOOTHIE COLORS
sheet = wb['REG_ADDONS']
fourk_cell = sheet.cell(1, 21)
threek_cell = sheet.cell(2, 21)
twok_cell = sheet.cell(3, 21)
onek_cell = sheet.cell(4, 21)
zerok_cell = sheet.cell(5, 21)
if fourk_cell.has_style:
    if 'FOURk_style' not in wb.named_styles:
        FOURk_style = NamedStyle(name='FOURk_style')
        FOURk_style.font = copy(fourk_cell.font)
        FOURk_style.border = copy(fourk_cell.border)
        FOURk_style.fill = copy(fourk_cell.fill)
        FOURk_style.number_format = copy(fourk_cell.number_format)
        FOURk_style.protection = copy(fourk_cell.protection)
        FOURk_style.alignment = copy(fourk_cell.alignment)
        wb.add_named_style(FOURk_style)
if threek_cell.has_style:
    if 'THREEk_style' not in wb.named_styles:
        THREEk_style = NamedStyle(name='THREEk_style')
        THREEk_style.font = copy(threek_cell.font)
        THREEk_style.border = copy(threek_cell.border)
        THREEk_style.fill = copy(threek_cell.fill)
        THREEk_style.number_format = copy(threek_cell.number_format)
        THREEk_style.protection = copy(threek_cell.protection)
        THREEk_style.alignment = copy(threek_cell.alignment)
        wb.add_named_style(THREEk_style)
if twok_cell.has_style:
    if 'TWOk_style' not in wb.named_styles:
        TWOk_style = NamedStyle(name='TWOk_style')
        TWOk_style.font = copy(twok_cell.font)
        TWOk_style.border = copy(twok_cell.border)
        TWOk_style.fill = copy(twok_cell.fill)
        TWOk_style.number_format = copy(twok_cell.number_format)
        TWOk_style.protection = copy(twok_cell.protection)
        TWOk_style.alignment = copy(twok_cell.alignment)
        wb.add_named_style(TWOk_style)
if onek_cell.has_style:
    if 'ONEk_style' not in wb.named_styles:
        ONEk_style = NamedStyle(name='ONEk_style')
        ONEk_style.font = copy(onek_cell.font)
        ONEk_style.border = copy(onek_cell.border)
        ONEk_style.fill = copy(onek_cell.fill)
        ONEk_style.number_format = copy(onek_cell.number_format)
        ONEk_style.protection = copy(onek_cell.protection)
        ONEk_style.alignment = copy(onek_cell.alignment)
        wb.add_named_style(ONEk_style)
if zerok_cell.has_style:
    if 'ZEROk_style' not in wb.named_styles:
        ZEROk_style = NamedStyle(name='ZEROk_style')
        ZEROk_style.font = copy(zerok_cell.font)
        ZEROk_style.border = copy(zerok_cell.border)
        ZEROk_style.fill = copy(zerok_cell.fill)
        ZEROk_style.number_format = copy(zerok_cell.number_format)
        ZEROk_style.protection = copy(zerok_cell.protection)
        ZEROk_style.alignment = copy(zerok_cell.alignment)
        wb.add_named_style(ZEROk_style)

kids_column = 2
for row in range(4, sheet.max_row + 1):
    kids = sheet.cell(row, kids_column)
    if kids.value == '4':
        for col in range(1, 4):
            Cell = sheet.cell(row, col)
            Cell.style = 'FOURk_style'
    elif kids.value == '3':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'THREEk_style'
    elif kids.value == '2':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'TWOk_style'
    elif kids.value == '1':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ONEk_style'
    elif kids.value == '0':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ZEROk_style'

for row in range(4, sheet.max_row + 1):
    for col in range(5, 12):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value = ''
    for col in range (13, 20):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value = ''

wb.save('000. DONOTUSE.xlsx')


###GF SMOOTHIES
reg_sheet = wb['ShipRep_GF']
list_sheet = wb['VIP.GF_ADDONS']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']

alpha_column = 26
name_column = 1
kids_column = 4
packlist_row = 22
smoothie_column = 6
antiox_column = 8
bigred_column = 9
blue_column = 10
boost_column = 11
green_column = 12
pbj_column = 13
vitc_column = 14
cookies_column = 15
break_column = 17
chip_column = 18
chipgf_column = 19
or_column = 20
orgf_column = 21
sugar_column = 22
sugargf_column = 23

for row in range(2, reg_sheet.max_row +1):
    alpha_cell = reg_sheet.cell(row, alpha_column)
    name_cell = reg_sheet.cell(row, name_column)
    kids_cell = reg_sheet.cell(row,kids_column)
    smoothie_cell = reg_sheet.cell(row,smoothie_column)
    antiox_cell = reg_sheet.cell(row, antiox_column)
    bigred_cell = reg_sheet.cell(row, bigred_column)
    blue_cell = reg_sheet.cell(row, blue_column)
    boost_cell = reg_sheet.cell(row, boost_column)
    green_cell = reg_sheet.cell(row, green_column)
    pbj_cell = reg_sheet.cell(row, pbj_column)
    vitc_cell = reg_sheet.cell(row,vitc_column)
    cookies_cell = reg_sheet.cell(row, cookies_column)
    break_cell = reg_sheet.cell(row, break_column)
    chip_cell = reg_sheet.cell(row, chip_column)
    chipgf_cell = reg_sheet.cell(row, chipgf_column)
    or_cell = reg_sheet.cell(row, or_column)
    orgf_cell = reg_sheet.cell(row, orgf_column)
    sugar_cell = reg_sheet.cell(row,sugar_column)
    sugargf_cell = reg_sheet.cell(row,sugargf_column)
    alpha = alpha_cell.value
    name = name_cell.value
    kids = kids_cell.value
    if alpha_cell.value in addon_list:
        pack_alpha = list_sheet.cell(packlist_row, 1)
        pack_kids = list_sheet.cell(packlist_row, 2)
        pack_name = list_sheet.cell(packlist_row, 3)
        pack_smoothies = list_sheet.cell(packlist_row, 4)
        pack_antiox = list_sheet.cell(packlist_row, 5)
        pack_bigred = list_sheet.cell(packlist_row, 6)
        pack_blue = list_sheet.cell(packlist_row, 7)
        pack_boost = list_sheet.cell(packlist_row, 8)
        pack_green = list_sheet.cell(packlist_row, 9)
        pack_pbj = list_sheet.cell(packlist_row, 10)
        pack_vitc = list_sheet.cell(packlist_row, 11)
        pack_cookie = list_sheet.cell(packlist_row, 12)
        pack_break = list_sheet.cell(packlist_row, 13)
        pack_chip = list_sheet.cell(packlist_row, 14)
        pack_chipgf = list_sheet.cell(packlist_row, 15)
        pack_or = list_sheet.cell(packlist_row, 16)
        pack_orgf = list_sheet.cell(packlist_row, 17)
        pack_sugar = list_sheet.cell(packlist_row, 18)
        pack_sugargf = list_sheet.cell(packlist_row, 19)
        pack_alpha.value = alpha
        pack_name.value = name
        pack_kids.value = kids
        pack_smoothies.value = smoothie_cell.value
        pack_antiox.value = antiox_cell.value
        pack_bigred.value = bigred_cell.value
        pack_blue.value = blue_cell.value
        pack_boost.value = boost_cell.value
        pack_green.value = green_cell.value
        pack_pbj.value = pbj_cell.value
        pack_vitc.value = vitc_cell.value
        pack_cookie.value = cookies_cell.value
        pack_break.value = break_cell.value
        pack_chip.value = chip_cell.value
        pack_chipgf.value = chipgf_cell.value
        pack_or.value = or_cell.value
        pack_orgf.value = orgf_cell.value
        pack_sugar.value = sugar_cell.value
        pack_sugargf.value = sugargf_cell.value
        packlist_row = packlist_row + 1
wb.save('000. DONOTUSE.xlsx')
wb = openpyxl.load_workbook('000. DONOTUSE.xlsx')
##GF SMOOTHIE COLORS
sheet = wb['VIP.GF_ADDONS']
fourk_cell = sheet.cell(1, 21)
threek_cell = sheet.cell(2, 21)
twok_cell = sheet.cell(3, 21)
onek_cell = sheet.cell(4, 21)
zerok_cell = sheet.cell(5, 21)
if fourk_cell.has_style:
    if 'FOURk_style' not in wb.named_styles:
        FOURk_style = NamedStyle(name='FOURk_style')
        FOURk_style.font = copy(fourk_cell.font)
        FOURk_style.border = copy(fourk_cell.border)
        FOURk_style.fill = copy(fourk_cell.fill)
        FOURk_style.number_format = copy(fourk_cell.number_format)
        FOURk_style.protection = copy(fourk_cell.protection)
        FOURk_style.alignment = copy(fourk_cell.alignment)
        wb.add_named_style(FOURk_style)
if threek_cell.has_style:
    if 'THREEk_style' not in wb.named_styles:
        THREEk_style = NamedStyle(name='THREEk_style')
        THREEk_style.font = copy(threek_cell.font)
        THREEk_style.border = copy(threek_cell.border)
        THREEk_style.fill = copy(threek_cell.fill)
        THREEk_style.number_format = copy(threek_cell.number_format)
        THREEk_style.protection = copy(threek_cell.protection)
        THREEk_style.alignment = copy(threek_cell.alignment)
        wb.add_named_style(THREEk_style)
if twok_cell.has_style:
    if 'TWOk_style' not in wb.named_styles:
        TWOk_style = NamedStyle(name='TWOk_style')
        TWOk_style.font = copy(twok_cell.font)
        TWOk_style.border = copy(twok_cell.border)
        TWOk_style.fill = copy(twok_cell.fill)
        TWOk_style.number_format = copy(twok_cell.number_format)
        TWOk_style.protection = copy(twok_cell.protection)
        TWOk_style.alignment = copy(twok_cell.alignment)
        wb.add_named_style(TWOk_style)
if onek_cell.has_style:
    if 'ONEk_style' not in wb.named_styles:
        ONEk_style = NamedStyle(name='ONEk_style')
        ONEk_style.font = copy(onek_cell.font)
        ONEk_style.border = copy(onek_cell.border)
        ONEk_style.fill = copy(onek_cell.fill)
        ONEk_style.number_format = copy(onek_cell.number_format)
        ONEk_style.protection = copy(onek_cell.protection)
        ONEk_style.alignment = copy(onek_cell.alignment)
        wb.add_named_style(ONEk_style)
if zerok_cell.has_style:
    if 'ZEROk_style' not in wb.named_styles:
        ZEROk_style = NamedStyle(name='ZEROk_style')
        ZEROk_style.font = copy(zerok_cell.font)
        ZEROk_style.border = copy(zerok_cell.border)
        ZEROk_style.fill = copy(zerok_cell.fill)
        ZEROk_style.number_format = copy(zerok_cell.number_format)
        ZEROk_style.protection = copy(zerok_cell.protection)
        ZEROk_style.alignment = copy(zerok_cell.alignment)
        wb.add_named_style(ZEROk_style)

kids_column = 2
for row in range(4, sheet.max_row + 1):
    kids = sheet.cell(row, kids_column)
    if kids.value == '4':
        for col in range(1, 4):
            Cell = sheet.cell(row, col)
            Cell.style = 'FOURk_style'
    elif kids.value == '3':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'THREEk_style'
    elif kids.value == '2':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'TWOk_style'
    elif kids.value == '1':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ONEk_style'
    elif kids.value == '0':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ZEROk_style'

for row in range(4, sheet.max_row + 1):
    for col in range(5, 12):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value = ''
    for col in range (13, 20):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value = ''

wb.save('000. DONOTUSE.xlsx')


### ADD DELIVERY ROUTES BACK IN
wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('000. DONOTUSE.xlsx')
final_sheet = fwb['ShipRep_GF']
sheet = wb['ShipRep_GF']
delivery_column = 25
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('000. DONOTUSE.xlsx')

wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('000. DONOTUSE.xlsx')
final_sheet = fwb['ShipRep_REG']
sheet = wb['ShipRep_REG']
delivery_column = 25
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('000. DONOTUSE.xlsx')

wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('000. DONOTUSE.xlsx')
final_sheet = fwb['ShipRep_VIP']
sheet = wb['ShipRep_VIP']
delivery_column = 26
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('6. Shipping Report Packlist OUTPUT.xlsx')