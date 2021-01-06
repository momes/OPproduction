import openpyxl as xl
meal_numbers_wb = xl.load_workbook('3. Label Numbers INPUT.xlsx')

##SHEET NAMES
count_sheet = meal_numbers_wb['COUNT']
red_sheet = meal_numbers_wb['RED']
teal_sheet = meal_numbers_wb['TEAL']
green_pot_sheet = meal_numbers_wb['GREEN']
yellow_sheet = meal_numbers_wb['YELLOW']
blue_pot_sheet = meal_numbers_wb['BLUE']
orange_pot_sheet = meal_numbers_wb['ORANGE']
liteblue_sheet = meal_numbers_wb['LITEBLUE']
grey_sheet = meal_numbers_wb['GREY']

##KID SIZE ROW VALUES
four_kids = 3
three_kids = 4
two_kids = 5
one_kid = 6
zero_kids = 7
four_kids_GF = 10
three_kids_GF = 11
two_kids_GF = 12
one_kid_GF = 13
zero_kids_GF = 14
##COLUMN values
red_column = 2
teal_column = 3
green_pot_column = 4
yellow_column = 5
blue_pot_column = 6
b_tacos_column = 7
t_tacos_column = 8
v_tacos_column = 9
liteblue_column = 10
b_burg_column = 11
t_burg_column = 12
v_burg_column = 13
orange_pot_column = 14
pep_pizza_column = 15
v_pizza_column = 16
grey_column = 17

use_by_date = 'Best if Used By: ' + input('Use By: ')
freeze_by_date = 'Use or Freeze By: ' + input('Freeze By: ')

###
###
###
######RED POT SHEET#######

###NUMBERS
fourk_red = count_sheet.cell(four_kids, red_column)
threek_red = count_sheet.cell(three_kids, red_column)
twok_red = count_sheet.cell(two_kids, red_column)
onek_red = count_sheet.cell(one_kid, red_column)
zerok_red = count_sheet.cell(zero_kids, red_column)
fourk_red_GF = count_sheet.cell(four_kids_GF, red_column)
threek_red_GF = count_sheet.cell(three_kids_GF, red_column)
twok_red_GF = count_sheet.cell(two_kids_GF, red_column)
onek_red_GF = count_sheet.cell(one_kid_GF, red_column)
zerok_red_GF = count_sheet.cell(zero_kids_GF, red_column)

regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_red.value + zerok_red_GF.value + 1
blue_range = pink_range + onek_red.value + onek_red_GF.value + 1
white_range = blue_range + twok_red.value + twok_red_GF.value + 1
green_range = white_range + threek_red.value + threek_red_GF.value + 1
orange_range = green_range + fourk_red.value + fourk_red_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = red_sheet.cell(row, 1)
    use_by_date_column = red_sheet.cell(row, 4)
    ##INGREDIENTS
    ingredient_column.value = 'Organic Ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = red_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = red_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = red_sheet.cell(row, 2)
    if row >= pink_range - zerok_red_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = red_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = red_sheet.cell(row, 2)
    if row >= blue_range - onek_red_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = red_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = red_sheet.cell(row, 2)
    if row >= white_range - twok_red_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = red_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = red_sheet.cell(row, 2)
    if row >= green_range - threek_red_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = red_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = red_sheet.cell(row, 2)
    if row >= orange_range - fourk_red_GF.value:
        gluten_free_column.value = gluten_free


###
###
###
######teal POT SHEET#######

###NUMBERS
fourk_teal = count_sheet.cell(four_kids, teal_column)
threek_teal = count_sheet.cell(three_kids, teal_column)
twok_teal = count_sheet.cell(two_kids, teal_column)
onek_teal = count_sheet.cell(one_kid, teal_column)
zerok_teal = count_sheet.cell(zero_kids, teal_column)
fourk_teal_GF = count_sheet.cell(four_kids_GF, teal_column)
threek_teal_GF = count_sheet.cell(three_kids_GF, teal_column)
twok_teal_GF = count_sheet.cell(two_kids_GF, teal_column)
onek_teal_GF = count_sheet.cell(one_kid_GF, teal_column)
zerok_teal_GF = count_sheet.cell(zero_kids_GF, teal_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_teal.value + zerok_teal_GF.value + 1
blue_range = pink_range + onek_teal.value + onek_teal_GF.value + 1
white_range = blue_range + twok_teal.value + twok_teal_GF.value + 1
green_range = white_range + threek_teal.value + threek_teal_GF.value + 1
orange_range = green_range + fourk_teal.value + fourk_teal_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = teal_sheet.cell(row, 1)
    use_by_date_column = teal_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = teal_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = teal_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = teal_sheet.cell(row, 2)
    if row >= pink_range - zerok_teal_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = teal_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = teal_sheet.cell(row, 2)
    if row >= blue_range - onek_teal_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = teal_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = teal_sheet.cell(row, 2)
    if row >= white_range - twok_teal_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = teal_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = teal_sheet.cell(row, 2)
    if row >= green_range - threek_teal_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = teal_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = teal_sheet.cell(row, 2)
    if row >= orange_range - fourk_teal_GF.value:
        gluten_free_column.value = gluten_free



###
###
###
######green_pot POT SHEET#######

###NUMBERS
fourk_green_pot = count_sheet.cell(four_kids, green_pot_column)
threek_green_pot = count_sheet.cell(three_kids, green_pot_column)
twok_green_pot = count_sheet.cell(two_kids, green_pot_column)
onek_green_pot = count_sheet.cell(one_kid, green_pot_column)
zerok_green_pot = count_sheet.cell(zero_kids, green_pot_column)
fourk_green_pot_GF = count_sheet.cell(four_kids_GF, green_pot_column)
threek_green_pot_GF = count_sheet.cell(three_kids_GF, green_pot_column)
twok_green_pot_GF = count_sheet.cell(two_kids_GF, green_pot_column)
onek_green_pot_GF = count_sheet.cell(one_kid_GF, green_pot_column)
zerok_green_pot_GF = count_sheet.cell(zero_kids_GF, green_pot_column)

regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_green_pot.value + zerok_green_pot_GF.value + 1
blue_range = pink_range + onek_green_pot.value + onek_green_pot_GF.value + 1
white_range = blue_range + twok_green_pot.value + twok_green_pot_GF.value + 1
green_range = white_range + threek_green_pot.value + threek_green_pot_GF.value + 1
orange_range = green_range + fourk_green_pot.value + fourk_green_pot_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = green_pot_sheet.cell(row, 1)
    use_by_date_column = green_pot_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = green_pot_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row >= pink_range - zerok_green_pot_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = green_pot_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row >= blue_range - onek_green_pot_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = green_pot_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row >= white_range - twok_green_pot_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = green_pot_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row >= green_range - threek_green_pot_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = green_pot_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = green_pot_sheet.cell(row, 2)
    if row >= orange_range - fourk_green_pot_GF.value:
        gluten_free_column.value = gluten_free

###
###
###
######yellow POT SHEET#######

###NUMBERS
fourk_yellow = count_sheet.cell(four_kids, yellow_column)
threek_yellow = count_sheet.cell(three_kids, yellow_column)
twok_yellow = count_sheet.cell(two_kids, yellow_column)
onek_yellow = count_sheet.cell(one_kid, yellow_column)
zerok_yellow = count_sheet.cell(zero_kids, yellow_column)
fourk_yellow_GF = count_sheet.cell(four_kids_GF, yellow_column)
threek_yellow_GF = count_sheet.cell(three_kids_GF, yellow_column)
twok_yellow_GF = count_sheet.cell(two_kids_GF, yellow_column)
onek_yellow_GF = count_sheet.cell(one_kid_GF, yellow_column)
zerok_yellow_GF = count_sheet.cell(zero_kids_GF, yellow_column)

regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_yellow.value + zerok_yellow_GF.value + 1
blue_range = pink_range + onek_yellow.value + onek_yellow_GF.value + 1
white_range = blue_range + twok_yellow.value + twok_yellow_GF.value + 1
green_range = white_range + threek_yellow.value + threek_yellow_GF.value + 1
orange_range = green_range + fourk_yellow.value + fourk_yellow_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = yellow_sheet.cell(row, 1)
    use_by_date_column = yellow_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = yellow_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row >= pink_range - zerok_yellow_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = yellow_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row >= blue_range - onek_yellow_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = yellow_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row >= white_range - twok_yellow_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = yellow_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row >= green_range - threek_yellow_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = yellow_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = yellow_sheet.cell(row, 2)
    if row >= orange_range - fourk_yellow_GF.value:
        gluten_free_column.value = gluten_free

###
###
###
######blue_pot POT SHEET#######

###NUMBERS
fourk_blue_pot = count_sheet.cell(four_kids, blue_pot_column)
threek_blue_pot = count_sheet.cell(three_kids, blue_pot_column)
twok_blue_pot = count_sheet.cell(two_kids, blue_pot_column)
onek_blue_pot = count_sheet.cell(one_kid, blue_pot_column)
zerok_blue_pot = count_sheet.cell(zero_kids, blue_pot_column)
fourk_blue_pot_GF = count_sheet.cell(four_kids_GF, blue_pot_column)
threek_blue_pot_GF = count_sheet.cell(three_kids_GF, blue_pot_column)
twok_blue_pot_GF = count_sheet.cell(two_kids_GF, blue_pot_column)
onek_blue_pot_GF = count_sheet.cell(one_kid_GF, blue_pot_column)
zerok_blue_pot_GF = count_sheet.cell(zero_kids_GF, blue_pot_column)

regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_blue_pot.value + zerok_blue_pot_GF.value + 1
blue_range = pink_range + onek_blue_pot.value + onek_blue_pot_GF.value + 1
white_range = blue_range + twok_blue_pot.value + twok_blue_pot_GF.value + 1
green_range = white_range + threek_blue_pot.value + threek_blue_pot_GF.value + 1
orange_range = green_range + fourk_blue_pot.value + fourk_blue_pot_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = blue_pot_sheet.cell(row, 1)
    use_by_date_column = blue_pot_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = blue_pot_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row >= pink_range - zerok_blue_pot_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = blue_pot_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row >= blue_range - onek_blue_pot_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = blue_pot_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row >= white_range - twok_blue_pot_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = blue_pot_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row >= green_range - threek_blue_pot_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = blue_pot_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = blue_pot_sheet.cell(row, 2)
    if row >= orange_range - fourk_blue_pot_GF.value:
        gluten_free_column.value = gluten_free

###
###
###
######orange_pot POT SHEET#######

###NUMBERS
fourk_orange_pot = count_sheet.cell(four_kids, orange_pot_column)
threek_orange_pot = count_sheet.cell(three_kids, orange_pot_column)
twok_orange_pot = count_sheet.cell(two_kids, orange_pot_column)
onek_orange_pot = count_sheet.cell(one_kid, orange_pot_column)
zerok_orange_pot = count_sheet.cell(zero_kids, orange_pot_column)
fourk_orange_pot_GF = count_sheet.cell(four_kids_GF, orange_pot_column)
threek_orange_pot_GF = count_sheet.cell(three_kids_GF, orange_pot_column)
twok_orange_pot_GF = count_sheet.cell(two_kids_GF, orange_pot_column)
onek_orange_pot_GF = count_sheet.cell(one_kid_GF, orange_pot_column)
zerok_orange_pot_GF = count_sheet.cell(zero_kids_GF, orange_pot_column)

regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_orange_pot.value + zerok_orange_pot_GF.value + 1
blue_range = pink_range + onek_orange_pot.value + onek_orange_pot_GF.value + 1
white_range = blue_range + twok_orange_pot.value + twok_orange_pot_GF.value + 1
green_range = white_range + threek_orange_pot.value + threek_orange_pot_GF.value + 1
orange_range = green_range + fourk_orange_pot.value + fourk_orange_pot_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = orange_pot_sheet.cell(row, 1)
    use_by_date_column = orange_pot_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = orange_pot_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row >= pink_range - zerok_orange_pot_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = orange_pot_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row >= blue_range - onek_orange_pot_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = orange_pot_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row >= white_range - twok_orange_pot_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = orange_pot_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row >= green_range - threek_orange_pot_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = orange_pot_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = orange_pot_sheet.cell(row, 2)
    if row >= orange_range - fourk_orange_pot_GF.value:
        gluten_free_column.value = gluten_free

###
###
###
######liteblue POT SHEET#######

###NUMBERS
fourk_liteblue = count_sheet.cell(four_kids, liteblue_column)
threek_liteblue = count_sheet.cell(three_kids, liteblue_column)
twok_liteblue = count_sheet.cell(two_kids, liteblue_column)
onek_liteblue = count_sheet.cell(one_kid, liteblue_column)
zerok_liteblue = count_sheet.cell(zero_kids, liteblue_column)
fourk_liteblue_GF = count_sheet.cell(four_kids_GF, liteblue_column)
threek_liteblue_GF = count_sheet.cell(three_kids_GF, liteblue_column)
twok_liteblue_GF = count_sheet.cell(two_kids_GF, liteblue_column)
onek_liteblue_GF = count_sheet.cell(one_kid_GF, liteblue_column)
zerok_liteblue_GF = count_sheet.cell(zero_kids_GF, liteblue_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_liteblue.value + zerok_liteblue_GF.value + 1
blue_range = pink_range + onek_liteblue.value + onek_liteblue_GF.value + 1
white_range = blue_range + twok_liteblue.value + twok_liteblue_GF.value + 1
green_range = white_range + threek_liteblue.value + threek_liteblue_GF.value + 1
orange_range = green_range + fourk_liteblue.value + fourk_liteblue_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = liteblue_sheet.cell(row, 1)
    use_by_date_column = liteblue_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = liteblue_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row >= pink_range - zerok_liteblue_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = liteblue_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row >= blue_range - onek_liteblue_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = liteblue_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row >= white_range - twok_liteblue_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = liteblue_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row >= green_range - threek_liteblue_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = liteblue_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = liteblue_sheet.cell(row, 2)
    if row >= orange_range - fourk_liteblue_GF.value:
        gluten_free_column.value = gluten_free


###
###
###
######grey POT SHEET#######

###NUMBERS
fourk_grey = count_sheet.cell(four_kids, grey_column)
threek_grey = count_sheet.cell(three_kids, grey_column)
twok_grey = count_sheet.cell(two_kids, grey_column)
onek_grey = count_sheet.cell(one_kid, grey_column)
zerok_grey = count_sheet.cell(zero_kids, grey_column)
fourk_grey_GF = count_sheet.cell(four_kids_GF, grey_column)
threek_grey_GF = count_sheet.cell(three_kids_GF, grey_column)
twok_grey_GF = count_sheet.cell(two_kids_GF, grey_column)
onek_grey_GF = count_sheet.cell(one_kid_GF, grey_column)
zerok_grey_GF = count_sheet.cell(zero_kids_GF, grey_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_grey.value + zerok_grey_GF.value + 1
blue_range = pink_range + onek_grey.value + onek_grey_GF.value + 1
white_range = blue_range + twok_grey.value + twok_grey_GF.value + 1
green_range = white_range + threek_grey.value + threek_grey_GF.value + 1
orange_range = green_range + fourk_grey.value + fourk_grey_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = grey_sheet.cell(row, 1)
    use_by_date_column = grey_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = grey_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = grey_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = grey_sheet.cell(row, 2)
    if row >= pink_range - zerok_grey_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = grey_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = grey_sheet.cell(row, 2)
    if row >= blue_range - onek_grey_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = grey_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = grey_sheet.cell(row, 2)
    if row >= white_range - twok_grey_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = grey_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = grey_sheet.cell(row, 2)
    if row >= green_range - threek_grey_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = grey_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = grey_sheet.cell(row, 2)
    if row >= orange_range - fourk_grey_GF.value:
        gluten_free_column.value = gluten_free



meal_numbers_wb.save('4. Label Numbers OUTPUT.xlsx')


####PROTEINS AND TOPPINGS
###
###
###
######TACOS#######
##SHEET NAMES
count_sheet = meal_numbers_wb['COUNT']
b_burg_sheet = meal_numbers_wb['B_BURG']
t_burg_sheet = meal_numbers_wb['T_BURG']
v_burg_sheet = meal_numbers_wb['V_BURG']
b_tacos_sheet = meal_numbers_wb['B_TACOS']
t_tacos_sheet = meal_numbers_wb['T_TACOS']
v_tacos_sheet = meal_numbers_wb['V_TACOS']
t_pizza_sheet = meal_numbers_wb['T_PIZZA']
v_pizza_sheet = meal_numbers_wb['V_PIZZA']

##KID SIZE ROW VALUES
four_kids = 3
three_kids = 4
two_kids = 5
one_kid = 6
zero_kids = 7
four_kids_GF = 10
three_kids_GF = 11
two_kids_GF = 12
one_kid_GF = 13
zero_kids_GF = 14
##COLUMN values
b_tacos_column = 7
t_tacos_column = 8
v_tacos_column = 9
b_burg_column = 11
t_burg_column = 12
v_burg_column = 13
t_pizza_column = 15
v_pizza_column = 16


###NUMBERS
fourk_b_tacos = count_sheet.cell(four_kids, b_tacos_column)
threek_b_tacos = count_sheet.cell(three_kids, b_tacos_column)
twok_b_tacos = count_sheet.cell(two_kids, b_tacos_column)
onek_b_tacos = count_sheet.cell(one_kid, b_tacos_column)
zerok_b_tacos = count_sheet.cell(zero_kids, b_tacos_column)
fourk_b_tacos_GF = count_sheet.cell(four_kids_GF, b_tacos_column)
threek_b_tacos_GF = count_sheet.cell(three_kids_GF, b_tacos_column)
twok_b_tacos_GF = count_sheet.cell(two_kids_GF, b_tacos_column)
onek_b_tacos_GF = count_sheet.cell(one_kid_GF, b_tacos_column)
zerok_b_tacos_GF = count_sheet.cell(zero_kids_GF, b_tacos_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_b_tacos.value + zerok_b_tacos_GF.value + 1
blue_range = pink_range + onek_b_tacos.value + onek_b_tacos_GF.value + 1
white_range = blue_range + twok_b_tacos.value + twok_b_tacos_GF.value + 1
green_range = white_range + threek_b_tacos.value + threek_b_tacos_GF.value + 1
orange_range = green_range + fourk_b_tacos.value + fourk_b_tacos_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = b_tacos_sheet.cell(row, 1)
    use_by_date_column = b_tacos_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic Ground Beef'
    use_by_date_column.value = freeze_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = b_tacos_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row >= pink_range - zerok_b_tacos_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = b_tacos_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row >= blue_range - onek_b_tacos_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = b_tacos_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row >= white_range - twok_b_tacos_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = b_tacos_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row >= green_range - threek_b_tacos_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = b_tacos_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = b_tacos_sheet.cell(row, 2)
    if row >= orange_range - fourk_b_tacos_GF.value:
        gluten_free_column.value = gluten_free

###
###
###
######t_tacos POT SHEET#######

###NUMBERS
fourk_t_tacos = count_sheet.cell(four_kids, t_tacos_column)
threek_t_tacos = count_sheet.cell(three_kids, t_tacos_column)
twok_t_tacos = count_sheet.cell(two_kids, t_tacos_column)
onek_t_tacos = count_sheet.cell(one_kid, t_tacos_column)
zerok_t_tacos = count_sheet.cell(zero_kids, t_tacos_column)
fourk_t_tacos_GF = count_sheet.cell(four_kids_GF, t_tacos_column)
threek_t_tacos_GF = count_sheet.cell(three_kids_GF, t_tacos_column)
twok_t_tacos_GF = count_sheet.cell(two_kids_GF, t_tacos_column)
onek_t_tacos_GF = count_sheet.cell(one_kid_GF, t_tacos_column)
zerok_t_tacos_GF = count_sheet.cell(zero_kids_GF, t_tacos_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_t_tacos.value + zerok_t_tacos_GF.value + 1
blue_range = pink_range + onek_t_tacos.value + onek_t_tacos_GF.value + 1
white_range = blue_range + twok_t_tacos.value + twok_t_tacos_GF.value + 1
green_range = white_range + threek_t_tacos.value + threek_t_tacos_GF.value + 1
orange_range = green_range + fourk_t_tacos.value + fourk_t_tacos_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = t_tacos_sheet.cell(row, 1)
    use_by_date_column = t_tacos_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic Ground Turkey'
    use_by_date_column.value = freeze_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = t_tacos_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row >= pink_range - zerok_t_tacos_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = t_tacos_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row >= blue_range - onek_t_tacos_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = t_tacos_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row >= white_range - twok_t_tacos_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = t_tacos_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row >= green_range - threek_t_tacos_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = t_tacos_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = t_tacos_sheet.cell(row, 2)
    if row >= orange_range - fourk_t_tacos_GF.value:
        gluten_free_column.value = gluten_free
        
###
###
###
######v_tacos POT SHEET#######

###NUMBERS
fourk_v_tacos = count_sheet.cell(four_kids, v_tacos_column)
threek_v_tacos = count_sheet.cell(three_kids, v_tacos_column)
twok_v_tacos = count_sheet.cell(two_kids, v_tacos_column)
onek_v_tacos = count_sheet.cell(one_kid, v_tacos_column)
zerok_v_tacos = count_sheet.cell(zero_kids, v_tacos_column)
fourk_v_tacos_GF = count_sheet.cell(four_kids_GF, v_tacos_column)
threek_v_tacos_GF = count_sheet.cell(three_kids_GF, v_tacos_column)
twok_v_tacos_GF = count_sheet.cell(two_kids_GF, v_tacos_column)
onek_v_tacos_GF = count_sheet.cell(one_kid_GF, v_tacos_column)
zerok_v_tacos_GF = count_sheet.cell(zero_kids_GF, v_tacos_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_v_tacos.value + zerok_v_tacos_GF.value + 1
blue_range = pink_range + onek_v_tacos.value + onek_v_tacos_GF.value + 1
white_range = blue_range + twok_v_tacos.value + twok_v_tacos_GF.value + 1
green_range = white_range + threek_v_tacos.value + threek_v_tacos_GF.value + 1
orange_range = green_range + fourk_v_tacos.value + fourk_v_tacos_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = v_tacos_sheet.cell(row, 1)
    use_by_date_column = v_tacos_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Beyond Meat Crumbles'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = v_tacos_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row >= pink_range - zerok_v_tacos_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = v_tacos_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row >= blue_range - onek_v_tacos_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = v_tacos_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row >= white_range - twok_v_tacos_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = v_tacos_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row >= green_range - threek_v_tacos_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = v_tacos_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = v_tacos_sheet.cell(row, 2)
    if row >= orange_range - fourk_v_tacos_GF.value:
        gluten_free_column.value = gluten_free
        
###
###
###
###BURGERS
######b_burg POT SHEET#######

###NUMBERS
fourk_b_burg = count_sheet.cell(four_kids, b_burg_column)
threek_b_burg = count_sheet.cell(three_kids, b_burg_column)
twok_b_burg = count_sheet.cell(two_kids, b_burg_column)
onek_b_burg = count_sheet.cell(one_kid, b_burg_column)
zerok_b_burg = count_sheet.cell(zero_kids, b_burg_column)
fourk_b_burg_GF = count_sheet.cell(four_kids_GF, b_burg_column)
threek_b_burg_GF = count_sheet.cell(three_kids_GF, b_burg_column)
twok_b_burg_GF = count_sheet.cell(two_kids_GF, b_burg_column)
onek_b_burg_GF = count_sheet.cell(one_kid_GF, b_burg_column)
zerok_b_burg_GF = count_sheet.cell(zero_kids_GF, b_burg_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_b_burg.value + zerok_b_burg_GF.value + 1
blue_range = pink_range + onek_b_burg.value + onek_b_burg_GF.value + 1
white_range = blue_range + twok_b_burg.value + twok_b_burg_GF.value + 1
green_range = white_range + threek_b_burg.value + threek_b_burg_GF.value + 1
orange_range = green_range + fourk_b_burg.value + fourk_b_burg_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = b_burg_sheet.cell(row, 1)
    use_by_date_column = b_burg_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic Grass-Fed Beef'
    use_by_date_column.value = freeze_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = b_burg_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row >= pink_range - zerok_b_burg_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = b_burg_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row >= blue_range - onek_b_burg_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = b_burg_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row >= white_range - twok_b_burg_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = b_burg_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row >= green_range - threek_b_burg_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = b_burg_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = b_burg_sheet.cell(row, 2)
    if row >= orange_range - fourk_b_burg_GF.value:
        gluten_free_column.value = gluten_free
        
###
###
###
######t_burg POT SHEET#######

###NUMBERS
fourk_t_burg = count_sheet.cell(four_kids, t_burg_column)
threek_t_burg = count_sheet.cell(three_kids, t_burg_column)
twok_t_burg = count_sheet.cell(two_kids, t_burg_column)
onek_t_burg = count_sheet.cell(one_kid, t_burg_column)
zerok_t_burg = count_sheet.cell(zero_kids, t_burg_column)
fourk_t_burg_GF = count_sheet.cell(four_kids_GF, t_burg_column)
threek_t_burg_GF = count_sheet.cell(three_kids_GF, t_burg_column)
twok_t_burg_GF = count_sheet.cell(two_kids_GF, t_burg_column)
onek_t_burg_GF = count_sheet.cell(one_kid_GF, t_burg_column)
zerok_t_burg_GF = count_sheet.cell(zero_kids_GF, t_burg_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_t_burg.value + zerok_t_burg_GF.value + 1
blue_range = pink_range + onek_t_burg.value + onek_t_burg_GF.value + 1
white_range = blue_range + twok_t_burg.value + twok_t_burg_GF.value + 1
green_range = white_range + threek_t_burg.value + threek_t_burg_GF.value + 1
orange_range = green_range + fourk_t_burg.value + fourk_t_burg_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = t_burg_sheet.cell(row, 1)
    use_by_date_column = t_burg_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic Ground Turkey'
    use_by_date_column.value = freeze_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = t_burg_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row >= pink_range - zerok_t_burg_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = t_burg_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row >= blue_range - onek_t_burg_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = t_burg_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row >= white_range - twok_t_burg_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = t_burg_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row >= green_range - threek_t_burg_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = t_burg_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = t_burg_sheet.cell(row, 2)
    if row >= orange_range - fourk_t_burg_GF.value:
        gluten_free_column.value = gluten_free
        
###
###
###
######v_burg POT SHEET#######

###NUMBERS
fourk_v_burg = count_sheet.cell(four_kids, v_burg_column)
threek_v_burg = count_sheet.cell(three_kids, v_burg_column)
twok_v_burg = count_sheet.cell(two_kids, v_burg_column)
onek_v_burg = count_sheet.cell(one_kid, v_burg_column)
zerok_v_burg = count_sheet.cell(zero_kids, v_burg_column)
fourk_v_burg_GF = count_sheet.cell(four_kids_GF, v_burg_column)
threek_v_burg_GF = count_sheet.cell(three_kids_GF, v_burg_column)
twok_v_burg_GF = count_sheet.cell(two_kids_GF, v_burg_column)
onek_v_burg_GF = count_sheet.cell(one_kid_GF, v_burg_column)
zerok_v_burg_GF = count_sheet.cell(zero_kids_GF, v_burg_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_v_burg.value + zerok_v_burg_GF.value + 1
blue_range = pink_range + onek_v_burg.value + onek_v_burg_GF.value + 1
white_range = blue_range + twok_v_burg.value + twok_v_burg_GF.value + 1
green_range = white_range + threek_v_burg.value + threek_v_burg_GF.value + 1
orange_range = green_range + fourk_v_burg.value + fourk_v_burg_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = v_burg_sheet.cell(row, 1)
    use_by_date_column = v_burg_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Impossible Burgers'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = v_burg_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row >= pink_range - zerok_v_burg_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = v_burg_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row >= blue_range - onek_v_burg_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = v_burg_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row >= white_range - twok_v_burg_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = v_burg_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row >= green_range - threek_v_burg_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = v_burg_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = v_burg_sheet.cell(row, 2)
    if row >= orange_range - fourk_v_burg_GF.value:
        gluten_free_column.value = gluten_free
        
###PIZZA
###
###
###
######t_pizza POT SHEET#######

###NUMBERS
fourk_t_pizza = count_sheet.cell(four_kids, t_pizza_column)
threek_t_pizza = count_sheet.cell(three_kids, t_pizza_column)
twok_t_pizza = count_sheet.cell(two_kids, t_pizza_column)
onek_t_pizza = count_sheet.cell(one_kid, t_pizza_column)
zerok_t_pizza = count_sheet.cell(zero_kids, t_pizza_column)
fourk_t_pizza_GF = count_sheet.cell(four_kids_GF, t_pizza_column)
threek_t_pizza_GF = count_sheet.cell(three_kids_GF, t_pizza_column)
twok_t_pizza_GF = count_sheet.cell(two_kids_GF, t_pizza_column)
onek_t_pizza_GF = count_sheet.cell(one_kid_GF, t_pizza_column)
zerok_t_pizza_GF = count_sheet.cell(zero_kids_GF, t_pizza_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_t_pizza.value + zerok_t_pizza_GF.value + 1
blue_range = pink_range + onek_t_pizza.value + onek_t_pizza_GF.value + 1
white_range = blue_range + twok_t_pizza.value + twok_t_pizza_GF.value + 1
green_range = white_range + threek_t_pizza.value + threek_t_pizza_GF.value + 1
orange_range = green_range + fourk_t_pizza.value + fourk_t_pizza_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = t_pizza_sheet.cell(row, 1)
    use_by_date_column = t_pizza_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic Turkey Pepperoni'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = t_pizza_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row >= pink_range - zerok_t_pizza_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = t_pizza_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row >= blue_range - onek_t_pizza_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = t_pizza_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row >= white_range - twok_t_pizza_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = t_pizza_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row >= green_range - threek_t_pizza_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = t_pizza_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = t_pizza_sheet.cell(row, 2)
    if row >= orange_range - fourk_t_pizza_GF.value:
        gluten_free_column.value = gluten_free
        
###
###
###
######v_pizza POT SHEET#######

###NUMBERS
fourk_v_pizza = count_sheet.cell(four_kids, v_pizza_column)
threek_v_pizza = count_sheet.cell(three_kids, v_pizza_column)
twok_v_pizza = count_sheet.cell(two_kids, v_pizza_column)
onek_v_pizza = count_sheet.cell(one_kid, v_pizza_column)
zerok_v_pizza = count_sheet.cell(zero_kids, v_pizza_column)
fourk_v_pizza_GF = count_sheet.cell(four_kids_GF, v_pizza_column)
threek_v_pizza_GF = count_sheet.cell(three_kids_GF, v_pizza_column)
twok_v_pizza_GF = count_sheet.cell(two_kids_GF, v_pizza_column)
onek_v_pizza_GF = count_sheet.cell(one_kid_GF, v_pizza_column)
zerok_v_pizza_GF = count_sheet.cell(zero_kids_GF, v_pizza_column)


regular = ''
gluten_free = 'Gluten Free'

##COLOR RANGES
pink_range = 17 + zerok_v_pizza.value + zerok_v_pizza_GF.value + 1
blue_range = pink_range + onek_v_pizza.value + onek_v_pizza_GF.value + 1
white_range = blue_range + twok_v_pizza.value + twok_v_pizza_GF.value + 1
green_range = white_range + threek_v_pizza.value + threek_v_pizza_GF.value + 1
orange_range = green_range + fourk_v_pizza.value + fourk_v_pizza_GF.value + 1

for row in range(2, orange_range):
    ingredient_column = v_pizza_sheet.cell(row, 1)
    use_by_date_column = v_pizza_sheet.cell(row, 4)
    ##ingredientS
    ingredient_column.value = 'Organic ingredient 1'
    use_by_date_column.value = use_by_date

###NO COLOR###
for row in range(2, 17):
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row <= 11:
        gluten_free_column.value = regular
    elif row >=12:
        gluten_free_column.value = gluten_free

###PINK###
for row in range(17, pink_range):
    pink_column = v_pizza_sheet.cell(row, 5)
    pink_column.value = '//////'
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row >= pink_range - zerok_v_pizza_GF.value:
        gluten_free_column.value = gluten_free

###BLUE###
for row in range(pink_range, blue_range):
    blue_column = v_pizza_sheet.cell(row, 6)
    blue_column.value = '//////'
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row >= blue_range - onek_v_pizza_GF.value:
        gluten_free_column.value = gluten_free

###WHITE###
for row in range(blue_range, white_range):
    white_column = v_pizza_sheet.cell(row, 7)
    white_column.value = '//////'
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row >= white_range - twok_v_pizza_GF.value:
        gluten_free_column.value = gluten_free

###GREEN###
for row in range(white_range, green_range):
    green_column = v_pizza_sheet.cell(row, 8)
    green_column.value = '//////'
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row >= green_range - threek_v_pizza_GF.value:
        gluten_free_column.value = gluten_free

###ORANGE###
for row in range(green_range, orange_range):
    orange_column = v_pizza_sheet.cell(row, 9)
    orange_column.value = '//////'
    gluten_free_column = v_pizza_sheet.cell(row, 2)
    if row >= orange_range - fourk_v_pizza_GF.value:
        gluten_free_column.value = gluten_free

meal_numbers_wb.save('4. Label Numbers OUTPUT.xlsx')